"""
openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
Com essa biblioteca será possível ler e escrever dados em células
específicas, formatar células, inserir gráficos,
criar fórmulas, adicionar imagens e outros elementos gráficos às suas
planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
Excel, como a criação de relatórios e análise de dados e/ou facilitando a
manipulação de grandes quantidades de informações.
Instalação necessária: pip install openpyxl
Documentação: https://openpyxl.readthedocs.io/en/stable/
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()

# Nome para a planilha
sheet_name = 'Minha planilha'
# Cria a planilha
workbook.create_sheet(sheet_name)
# Seleciona a planilha
worksheet: Worksheet = workbook[sheet_name]  # type: ignore
print(workbook.sheetnames)

# Remove sheet da planilha
workbook.remove(workbook['Sheet'])
print(workbook.sheetnames)

# Criando cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

estudantes = [
    # nome    idade   nota
    ['Victor',  14,   5.5],
    ['Marieli', 13,   9.7],
    ['Luiz',    15,   8.8],
    ['Judite',  16,   10],
]

# for i, linha_estudante in enumerate(estudantes, start=2):
#     for j, coluna_estudante in enumerate(linha_estudante, start=1):
#         worksheet.cell(i, j, coluna_estudante)

for e in estudantes:
    worksheet.append(e)

workbook.save(WORKBOOK_PATH)
print('')
# --------------------------------------------------------------------------------------------------------------------------------

"""
openpyxl - ler e alterar dados de uma planilha
"""

# Carregando um arquivo xlsx
workbook1: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name1 = 'Minha planilha'

worksheet1: Worksheet = workbook1[sheet_name1]

for row in worksheet1.iter_rows(min_row=2):
    for variavel in row:
        print(variavel.value, end='\t')

        if variavel.value == 'Marieli':
            worksheet1.cell(variavel.row, 2, 23)  # type: ignore
    print()

print(worksheet1['B3'].value)
# worksheet1['B3'].value = 14
# print(worksheet1['B3'].value)

workbook1.save(WORKBOOK_PATH)
