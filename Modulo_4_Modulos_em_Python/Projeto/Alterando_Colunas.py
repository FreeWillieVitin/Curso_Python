from os import path
from pathlib import Path
from jupyterlab_server import WorkspaceListApp
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WORK_PATH = 'C:\\Users\\HSVP\\Desktop\\Unimed maior pg15-01-24.xlsx'

work = Workbook()

planilha: Workbook = load_workbook(WORK_PATH)

sheet_name = 'Analí\xadtico'

dados: Worksheet = planilha[sheet_name]

for row, coluna in dados.iter_rows(min_row=1, max_col=4):
    print(coluna)
